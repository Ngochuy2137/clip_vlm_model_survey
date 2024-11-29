import torch
import clip
from PIL import Image as PILImage
import gc
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as openpyxlImage
import os

from python_utils.python_utils.printer import Printer

DEVICE = "cuda"

'''
Available models:
    print(clip.available_models())  
    ->  'RN50' 
        'RN101' 
        'RN50x4' 
        'RN50x16' 
        'RN50x64' 
        'ViT-B/32' 
        'ViT-B/16' 
        'ViT-L/14' 
        'ViT-L/14@336px'
'''

utils_printer = Printer()

class ClipTester:
    def __init__(self, object_name, device:str):
        self.device = device
        self.object_name = object_name
        
    def model_evaluation(self, text_list, image_path):
        self.text_list = text_list
        self.model_list = clip.available_models()
        self.image_path = image_path

        all_result = {}
        for model_name in clip.available_models():
            utils_printer.print_blue(f"Model: {model_name}", background=True)
            all_result[model_name] = self.model_run(model_name)
        
        self.save_to_excel(all_result, text_list, self.object_name, self.image_path)

    def model_run(self, model_name:str):
        if model_name not in self.model_list:
            utils_printer.print_red(f"Model {model_name} is not available")
            return
        
        model, preprocess = clip.load(model_name, device=self.device)
        image = preprocess(PILImage.open(self.image_path)).unsqueeze(0).to(self.device)
        text = clip.tokenize(self.text_list).to(self.device)

        with torch.no_grad():
            image_features = model.encode_image(image)
            text_features = model.encode_text(text)
            
            logits_per_image, logits_per_text = model(image, text)
            probs = logits_per_image.softmax(dim=-1).cpu().numpy()
            print("     Label probs:", probs)  # prints: [[0.9927937  0.00421068 0.00299572]]

            for text_i, prob_i in zip(self.text_list, probs[0]):
                print(f"        Text: {text_i} - Prob: {prob_i.max()*100:.3f} %")

        # Dọn dẹp mô hình sau khi hoàn thành vòng lặp
        del model
        torch.cuda.empty_cache()
        gc.collect()
        return probs[0]

    def save_to_excel(self, probs_data, text_list, object_name, image_path=None):
        # Tạo DataFrame từ dữ liệu xác suất
        text_list_with_quotes = [f'"{text}"' for text in text_list]  # Thêm dấu ""
        df = pd.DataFrame(probs_data, index=text_list_with_quotes)

        # Lấy đường dẫn folder của script
        folder_name = os.path.abspath(__file__)
        folder_name = folder_name.replace('clip_tester.py', 'results')
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        # Đường dẫn file Excel
        file_path = os.path.join(folder_name, 'clip_results.xlsx')
        sheet_name = f'Image of {object_name} test'

        # Kiểm tra nếu file đã tồn tại
        if os.path.exists(file_path):
            # Nếu file tồn tại, mở file và thêm sheet mới
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet_name)
        else:
            # Nếu file chưa tồn tại, tạo file mới
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=f'Image of {object_name} test')

        # Lưu ảnh nếu được cung cấp
        if image_path:
            self.add_image_to_excel(file_path, sheet_name, image_path, "B10")  # Thêm ảnh tại vị trí B10

        # In thông báo thành công
        print(f"Results are saved to {file_path}")

    def add_image_to_excel(self, file_path, sheet_name, image_path, cell_location):
        # Resize ảnh bằng Pillow
        img = PILImage.open(image_path)
        # get width and height
        width, height = img.size
        new_width = 200
        new_height = int(height * new_width / width)
        img = img.resize((new_width, new_height))  # Resize ảnh về kích thước mong muốn
        resized_image_path = "resized_image.png"
        img.save(resized_image_path)  # Lưu ảnh đã resize

        # Mở file Excel và thêm ảnh
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        # Thêm ảnh đã resize vào Excel
        img = openpyxlImage(resized_image_path)
        sheet.add_image(img, cell_location)

        workbook.save(file_path)
        print(f"Resized image added to '{file_path}' in sheet '{sheet_name}' at '{cell_location}'")


def main():
    # text_list = ["a diagram", "a plane", "a toy plane", "a real plane", "a iron plane", "a foam plane"]
    text_list = ["a plane", "a frisbee", "a boomerang"]
    image_path_dict = {
        'plane': 'images_object/big_plane.jpeg',
        'frisbee': 'images_object/frisbee_blue.jpeg',
        'boomerang': 'images_object/boomerang.jpeg'
    }

    print('Start testing')
    for object_name, image_path in image_path_dict.items():
        clip_tester = ClipTester(object_name, DEVICE)
        clip_tester.model_evaluation(text_list, image_path)

if __name__ == "__main__":
    main()
