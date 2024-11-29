import torch
import clip
from PIL import Image as PILImage
import gc
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as openpyxlImage
import os

from python_utils.python_utils.printer import Printer

DEVICE = "cuda"

utils_printer = Printer()

class ClipTester:
    def __init__(self, device:str):
        self.device = device
        self.available_models = clip.available_models()
        
    def model_evaluation(self, text_list, image_path_dict, model_list):
        label_img_list = list(image_path_dict.keys())
        image_path_list = list(image_path_dict.values())
        # check if len of text_list, image_path_list and label_img_list are the same
        if len(text_list) != len(image_path_list) or len(text_list) != len(label_img_list):
            utils_printer.print_red("Length of text_list, image_path_list and label_img_list must be the same")
            return
        
        all_result = {}
        for model_name in model_list:
            if model_name not in clip.available_models():
                utils_printer.print_red(f"Model {model_name} is not available")
                continue
            else:
                utils_printer.print_blue(f"Model: {model_name}", background=True)
                result_probs = self.model_run(text_list, image_path_list, model_name)
                if model_name not in all_result:
                    all_result[model_name] = {}
                for label_img, probs_text in zip(label_img_list, result_probs):
                    utils_printer.print_blue(f"    Image of object {label_img}:")
                    if label_img not in all_result[model_name]:
                        all_result[model_name][label_img] = {}
                    for text_i, prob_i in zip(text_list, probs_text):
                        all_result[model_name][label_img][text_i] = prob_i
                        print(f"        Label: {text_i} - Prob: {prob_i*100:.3f} %")

        
        # self.save_to_excel(all_result, text_list, self.object_name, self.image_path)

    def model_run(self, text_list, image_path_list, model_name:str):
        if model_name not in self.available_models:
            utils_printer.print_red(f"Model {model_name} is not available")
            return
        
        model, preprocess = clip.load(model_name, device=self.device)

        # load many images
        images = torch.stack([preprocess(PILImage.open(image_path)).to(self.device) for image_path in image_path_list])
        # image = preprocess(PILImage.open(self.image_path)).unsqueeze(0).to(self.device)
        texts = clip.tokenize(text_list).to(self.device)

        with torch.no_grad():
            image_features = model.encode_image(images)
            text_features = model.encode_text(texts)
            
            logits_per_image, logits_per_text = model(images, texts)
            probs = logits_per_image.softmax(dim=-1).cpu().numpy()
            print("     Label probs:", probs)  # prints: [[0.9927937  0.00421068 0.00299572]]

            for text_i, prob_i in zip(text_list, probs[0]):
                print(f"        Label: {text_i} - Prob: {prob_i.max()*100:.3f} %")

        # Dọn dẹp mô hình sau khi hoàn thành vòng lặp
        del model
        torch.cuda.empty_cache()
        gc.collect()
        return probs

    def save_to_excel(self, probs_data, text_list, object_name, image_path=None):
        # Tạo DataFrame từ dữ liệu xác suất
        df = pd.DataFrame(probs_data, index=text_list)

        # Lấy đường dẫn folder của script
        folder_name = os.path.abspath(__file__)
        folder_name = folder_name.replace('clip_tester.py', 'results')
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        # Đường dẫn file Excel
        file_path = os.path.join(folder_name, 'clip_results.xlsx')
        sheet_name = f'Probabilities of {object_name}'

        # Kiểm tra nếu file đã tồn tại
        if os.path.exists(file_path):
            # Nếu file tồn tại, mở file và thêm sheet mới
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet_name)
        else:
            # Nếu file chưa tồn tại, tạo file mới
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=f'Probabilities of {object_name}')

        # Lưu ảnh nếu được cung cấp
        if image_path:
            self.add_image_to_excel(file_path, sheet_name, image_path, "B10")  # Thêm ảnh tại vị trí B10

        # In thông báo thành công
        print(f"Results are saved to {file_path}")

    def add_image_to_excel(self, file_path, sheet_name, image_path, cell_location):
        # Resize ảnh bằng Pillow
        img = PILImage.open(image_path)
        # get width and height
        width, height = img.size
        new_width = 200
        new_height = int(height * new_width / width)
        img = img.resize((new_width, new_height))  # Resize ảnh về kích thước mong muốn
        resized_image_path = "resized_image.png"
        img.save(resized_image_path)  # Lưu ảnh đã resize

        # Mở file Excel và thêm ảnh
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        # Thêm ảnh đã resize vào Excel
        img = openpyxlImage(resized_image_path)
        sheet.add_image(img, cell_location)

        workbook.save(file_path)
        print(f"Resized image added to '{file_path}' in sheet '{sheet_name}' at '{cell_location}'")


def main():
    '''
    Available models:
        print(clip.available_models())  
        ->  'RN50' 
            'RN101' 
            'RN50x4' 
            'RN50x16' 
            'RN50x64' 
            'ViT-B/32' 
            'ViT-B/16' 
            'ViT-L/14' 
            'ViT-L/14@336px'
    '''
    # text_list = ["a diagram", "a plane", "a toy plane", "a real plane", "a iron plane", "a foam plane"]
    text_list = ["a plane", "a frisbee", "a boomerang"]
    image_path_dict = {
        'plane': 'images_object/big_plane.jpeg',
        'frisbee': 'images_object/frisbee_blue.jpeg',
        'boomerang': 'images_object/boomerang.jpeg'
    }
    model_list = ['RN50x16', 'RN50x64', 'ViT-L/14@336px']

    print('Start testing')
    clip_tester = ClipTester(DEVICE)
    clip_tester.model_evaluation(text_list, image_path_dict, model_list)

if __name__ == "__main__":
    main()
